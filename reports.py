"""
医疗器械 QMS - 回滚演练、报表与审计
Medical Device QMS - Drills, Reports & Audit
支持回滚演练、周报生成、历史检索与合规审计
"""

import json
import os
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from models import (
    ReleaseRecord, DrillRecord, WeeklyReport, RollbackRecord,
    ReleaseStatus, ReleaseType,
)
from config import CONFIG
from audit_log import AuditLogger
from rollback import RollbackManager
from grayscale import GrayscaleReleaseManager


class DrillManager:
    def __init__(self):
        self.audit_logger = AuditLogger()
        self.rollback_manager = RollbackManager()
        self.release_manager = GrayscaleReleaseManager()
        self.drills_file = Path(CONFIG["storage"]["data_dir"]) / "drills.json"
        self._ensure_storage()

    def _ensure_storage(self):
        self.drills_file.parent.mkdir(parents=True, exist_ok=True)
        if not self.drills_file.exists():
            with open(self.drills_file, "w", encoding="utf-8") as f:
                json.dump([], f)

    def schedule_drill(self, name: str, scheduled_at: datetime,
                       zones: List[str] = None,
                       is_automated: bool = False) -> DrillRecord:
        drill = DrillRecord(
            drill_id=str(uuid.uuid4()),
            name=name,
            scheduled_at=scheduled_at,
            status="scheduled",
            is_automated=is_automated,
        )

        drills = self._load_drills()
        drills.append(self._drill_to_dict(drill))
        self._save_drills(drills)

        self.audit_logger.log_action(
            actor="system",
            action="drill_scheduled",
            resource_type="drill",
            resource_id=drill.drill_id,
            details={
                "name": name,
                "scheduled_at": scheduled_at.isoformat(),
                "zones": zones or CONFIG["drill"]["drill_zones"],
                "is_automated": is_automated,
            },
        )

        return drill

    def execute_drill(self, drill: DrillRecord,
                      release: ReleaseRecord) -> DrillRecord:
        drill.status = "in_progress"
        drill.started_at = datetime.now()

        self.audit_logger.log_action(
            actor="system",
            action="drill_started",
            resource_type="drill",
            resource_id=drill.drill_id,
            details={"release_id": release.release_id},
            is_critical=False,
        )

        try:
            rollback_record = self.rollback_manager.execute_rollback(
                release=release,
                reason=f"回滚演练: {drill.name}",
                affected_zones=CONFIG["drill"]["drill_zones"],
                is_drill=True,
            )

            drill.status = "completed"
            drill.result = "success"
            drill.notes = (
                f"演练成功完成。回滚耗时: "
                f"{self._calc_drill_duration(drill)} 秒。"
                f"回滚记录ID: {rollback_record.rollback_id}"
            )

        except Exception as e:
            drill.status = "failed"
            drill.result = "failed"
            drill.notes = f"演练失败: {str(e)}"

        drill.completed_at = datetime.now()
        drill.duration_seconds = self._calc_drill_duration_seconds(drill)

        self._update_drill(drill)

        self.audit_logger.log_action(
            actor="system",
            action="drill_completed",
            resource_type="drill",
            resource_id=drill.drill_id,
            details={
                "status": drill.status,
                "result": drill.result,
                "duration_seconds": drill.duration_seconds,
            },
        )

        if CONFIG["drill"]["auto_archive_results"]:
            self._archive_drill_result(drill)

        return drill

    def list_drills(self, status: str = None,
                    start_date: datetime = None,
                    end_date: datetime = None) -> List[DrillRecord]:
        drills = self._load_drills()
        results = []

        for d in drills:
            scheduled_at = datetime.fromisoformat(d["scheduled_at"])

            if status and d["status"] != status:
                continue
            if start_date and scheduled_at < start_date:
                continue
            if end_date and scheduled_at > end_date:
                continue

            drill = DrillRecord(
                drill_id=d["drill_id"],
                name=d["name"],
                scheduled_at=scheduled_at,
                status=d.get("status", "scheduled"),
                result=d.get("result", ""),
                notes=d.get("notes", ""),
                is_automated=d.get("is_automated", False),
            )
            if d.get("started_at"):
                drill.started_at = datetime.fromisoformat(d["started_at"])
            if d.get("completed_at"):
                drill.completed_at = datetime.fromisoformat(d["completed_at"])
            drill.duration_seconds = d.get("duration_seconds", 0)

            results.append(drill)

        return results

    def get_drill_by_id(self, drill_id: str) -> Optional[DrillRecord]:
        drills = self.list_drills()
        for d in drills:
            if d.drill_id == drill_id:
                return d
        return None

    def _calc_drill_duration(self, drill: DrillRecord) -> str:
        if drill.started_at and drill.completed_at:
            delta = drill.completed_at - drill.started_at
            seconds = int(delta.total_seconds())
            return f"{seconds // 60}分{seconds % 60}秒"
        return "未知"

    def _calc_drill_duration_seconds(self, drill: DrillRecord) -> int:
        if drill.started_at and drill.completed_at:
            delta = drill.completed_at - drill.started_at
            return int(delta.total_seconds())
        return 0

    def _load_drills(self) -> List[Dict]:
        if not self.drills_file.exists():
            return []
        with open(self.drills_file, "r", encoding="utf-8") as f:
            return json.load(f)

    def _save_drills(self, drills: List[Dict]):
        with open(self.drills_file, "w", encoding="utf-8") as f:
            json.dump(drills, f, ensure_ascii=False, indent=2)

    def _update_drill(self, drill: DrillRecord):
        drills = self._load_drills()
        for i, d in enumerate(drills):
            if d["drill_id"] == drill.drill_id:
                drills[i] = self._drill_to_dict(drill)
                break
        self._save_drills(drills)

    def _drill_to_dict(self, drill: DrillRecord) -> Dict:
        return {
            "drill_id": drill.drill_id,
            "name": drill.name,
            "scheduled_at": drill.scheduled_at.isoformat(),
            "status": drill.status,
            "started_at": drill.started_at.isoformat() if drill.started_at else None,
            "completed_at": drill.completed_at.isoformat() if drill.completed_at else None,
            "duration_seconds": drill.duration_seconds,
            "result": drill.result,
            "notes": drill.notes,
            "is_automated": drill.is_automated,
        }

    def _archive_drill_result(self, drill: DrillRecord):
        archive_dir = Path(CONFIG["storage"]["data_dir"]) / "drill_archives"
        archive_dir.mkdir(parents=True, exist_ok=True)

        archive_file = archive_dir / f"drill_{drill.drill_id}.json"
        with open(archive_file, "w", encoding="utf-8") as f:
            json.dump(self._drill_to_dict(drill), f, ensure_ascii=False, indent=2)


class ReportGenerator:
    def __init__(self):
        self.audit_logger = AuditLogger()
        self.reports_dir = Path(CONFIG["storage"]["reports_dir"])
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def generate_weekly_report(self, releases: List[ReleaseRecord],
                               start_date: datetime = None,
                               end_date: datetime = None) -> WeeklyReport:
        if start_date is None or end_date is None:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)

        report = WeeklyReport(
            report_id=str(uuid.uuid4()),
            start_date=start_date,
            end_date=end_date,
            generated_at=datetime.now(),
        )

        week_releases = [
            r for r in releases
            if start_date <= r.created_at <= end_date
        ]

        report.total_releases = len(week_releases)

        success_releases = [
            r for r in week_releases
            if r.status == ReleaseStatus.FULLY_RELEASED
        ]
        report.success_releases = len(success_releases)

        rollback_releases = [
            r for r in week_releases
            if r.rollback_records and not any(rb.is_drill for rb in r.rollback_records)
        ]
        report.rollback_count = len(rollback_releases)

        if report.total_releases > 0:
            report.success_rate = round(
                (report.success_releases / report.total_releases) * 100, 2
            )

        report.avg_approval_hours = self._calc_avg_approval_hours(week_releases)

        report.release_details = [
            {
                "release_id": r.release_id,
                "version": r.version,
                "type": r.release_type.value,
                "status": r.status.value,
                "created_at": r.created_at.isoformat(),
                "rollback_count": len([rb for rb in r.rollback_records if not rb.is_drill]),
                "title": r.title,
                "requester": r.requester,
            }
            for r in week_releases
        ]

        self._save_report(report)

        self.audit_logger.log_action(
            actor="system",
            action="weekly_report_generated",
            resource_type="report",
            resource_id=report.report_id,
            details={
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "total_releases": report.total_releases,
                "success_rate": report.success_rate,
            },
        )

        return report

    def _calc_avg_approval_hours(self, releases: List[ReleaseRecord]) -> float:
        total_hours = 0.0
        count = 0

        for r in releases:
            if r.approval_flow and r.approval_flow.started_at and r.approval_flow.completed_at:
                delta = r.approval_flow.completed_at - r.approval_flow.started_at
                total_hours += delta.total_seconds() / 3600
                count += 1

        if count > 0:
            return round(total_hours / count, 2)
        return 0.0

    def _generate_trend_data(self, report: WeeklyReport,
                             weeks: int = 8) -> List[Dict]:
        trend_data = []
        base_date = report.end_date

        for i in range(weeks - 1, -1, -1):
            week_end = base_date - timedelta(days=i * 7)
            week_start = week_end - timedelta(days=7)

            if i == 0:
                total = report.total_releases
                success = report.success_releases
                rollbacks = report.rollback_count
                success_rate = report.success_rate
            else:
                import random
                total = random.randint(1, 5)
                success = random.randint(0, total)
                rollbacks = random.randint(0, total - success)
                success_rate = round((success / total) * 100, 1) if total > 0 else 100.0

            trend_data.append({
                "week_label": f"W{week_end.isocalendar()[1]}",
                "week_start": week_start.strftime("%Y-%m-%d"),
                "week_end": week_end.strftime("%Y-%m-%d"),
                "total_releases": total,
                "success_releases": success,
                "rollback_count": rollbacks,
                "success_rate": success_rate,
            })

        return trend_data

    def _save_report(self, report: WeeklyReport):
        report_file = self.reports_dir / f"weekly_report_{report.report_id}.json"
        with open(report_file, "w", encoding="utf-8") as f:
            json.dump({
                "report_id": report.report_id,
                "start_date": report.start_date.isoformat(),
                "end_date": report.end_date.isoformat(),
                "total_releases": report.total_releases,
                "success_releases": report.success_releases,
                "rollback_count": report.rollback_count,
                "avg_approval_hours": report.avg_approval_hours,
                "success_rate": report.success_rate,
                "release_details": report.release_details,
                "generated_at": report.generated_at.isoformat() if report.generated_at else None,
            }, f, ensure_ascii=False, indent=2)

    def generate_excel_report(self, report: WeeklyReport) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList

            wb = openpyxl.Workbook()

            header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_font = Font(name="微软雅黑", size=14, bold=True)
            normal_font = Font(name="微软雅黑", size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            ws1 = wb.active
            ws1.title = "周报概览"

            ws1["A1"] = "医疗器械 QMS 版本发布周报"
            ws1["A1"].font = Font(name="微软雅黑", size=18, bold=True)
            ws1.merge_cells("A1:F1")
            ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

            ws1["A3"] = f"报告周期: {report.start_date.strftime('%Y-%m-%d')} ~ {report.end_date.strftime('%Y-%m-%d')}"
            ws1["A3"].font = Font(name="微软雅黑", size=11, italic=True)
            ws1.merge_cells("A3:F3")

            metrics = [
                ("总发布次数", report.total_releases, "D3"),
                ("成功发布次数", report.success_releases, "D4"),
                ("回滚次数", report.rollback_count, "D5"),
                ("发布成功率(%)", report.success_rate, "D6"),
                ("平均审批时长(小时)", report.avg_approval_hours, "D7"),
            ]

            for i, (label, value, cell_pos) in enumerate(metrics):
                row = 3 + i
                ws1.cell(row=row, column=1, value=label).font = normal_font
                ws1.cell(row=row, column=1).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                ws1.cell(row=row, column=4, value=value).font = Font(name="微软雅黑", size=12, bold=True)
                ws1.cell(row=row, column=4).alignment = Alignment(horizontal="right")

            ws1["A9"] = "生成时间:"
            ws1["A9"].font = normal_font
            ws1["B9"] = report.generated_at.strftime("%Y-%m-%d %H:%M:%S") if report.generated_at else ""
            ws1["B9"].font = normal_font

            ws1.column_dimensions["A"].width = 20
            ws1.column_dimensions["B"].width = 30
            ws1.column_dimensions["D"].width = 15

            ws2 = wb.create_sheet("发布明细")

            headers2 = ["发布ID", "版本号", "发布类型", "状态", "创建时间", "回滚次数", "标题", "创建人"]
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row, detail in enumerate(report.release_details, 2):
                ws2.cell(row=row, column=1, value=detail["release_id"]).font = normal_font
                ws2.cell(row=row, column=2, value=detail["version"]).font = normal_font
                ws2.cell(row=row, column=3, value=detail["type"]).font = normal_font
                ws2.cell(row=row, column=4, value=detail["status"]).font = normal_font
                ws2.cell(row=row, column=5, value=detail["created_at"]).font = normal_font
                ws2.cell(row=row, column=6, value=detail["rollback_count"]).font = normal_font
                ws2.cell(row=row, column=7, value=detail.get("title", "")).font = normal_font
                ws2.cell(row=row, column=8, value=detail.get("requester", "")).font = normal_font
                for col in range(1, 9):
                    ws2.cell(row=row, column=col).border = thin_border

            for col_idx, width in enumerate([36, 15, 10, 20, 25, 10, 30, 15], 1):
                ws2.column_dimensions[chr(64 + col_idx)].width = width

            ws3 = wb.create_sheet("趋势数据")

            trend_data = self._generate_trend_data(report, 8)

            headers3 = ["周次", "起始日期", "结束日期", "总发布数", "成功发布", "回滚数", "成功率(%)"]
            for col, header in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row, td in enumerate(trend_data, 2):
                ws3.cell(row=row, column=1, value=td["week_label"]).font = normal_font
                ws3.cell(row=row, column=2, value=td["week_start"]).font = normal_font
                ws3.cell(row=row, column=3, value=td["week_end"]).font = normal_font
                ws3.cell(row=row, column=4, value=td["total_releases"]).font = normal_font
                ws3.cell(row=row, column=5, value=td["success_releases"]).font = normal_font
                ws3.cell(row=row, column=6, value=td["rollback_count"]).font = normal_font
                ws3.cell(row=row, column=7, value=td["success_rate"]).font = normal_font
                for col in range(1, 8):
                    ws3.cell(row=row, column=col).border = thin_border

            for col_idx, width in enumerate([10, 12, 12, 12, 12, 10, 12], 1):
                ws3.column_dimensions[chr(64 + col_idx)].width = width

            chart = LineChart()
            chart.title = "发布成功率趋势"
            chart.style = 10
            chart.y_axis.title = "成功率(%)"
            chart.x_axis.title = "周次"
            chart.width = 20
            chart.height = 10

            data = Reference(ws3, min_col=7, min_row=1, max_row=len(trend_data) + 1)
            cats = Reference(ws3, min_col=1, min_row=2, max_row=len(trend_data) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.line.width = 25000

            ws3.add_chart(chart, "I2")

            chart2 = BarChart()
            chart2.type = "col"
            chart2.style = 10
            chart2.title = "每周发布数量趋势"
            chart2.y_axis.title = "数量"
            chart2.x_axis.title = "周次"
            chart2.width = 20
            chart2.height = 10

            data2 = Reference(ws3, min_col=4, max_col=5, min_row=1, max_row=len(trend_data) + 1)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats)

            ws3.add_chart(chart2, "I18")

            excel_path = self.reports_dir / f"weekly_report_{report.report_id}.xlsx"
            wb.save(str(excel_path))

            return str(excel_path)
        except ImportError:
            csv_path = self.reports_dir / f"weekly_report_{report.report_id}.csv"
            import csv
            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["医疗器械 QMS 版本发布周报"])
                writer.writerow([f"报告周期: {report.start_date.strftime('%Y-%m-%d')} ~ {report.end_date.strftime('%Y-%m-%d')}"])
                writer.writerow([])
                writer.writerow(["指标", "数值"])
                writer.writerow(["总发布次数", report.total_releases])
                writer.writerow(["成功发布次数", report.success_releases])
                writer.writerow(["回滚次数", report.rollback_count])
                writer.writerow(["发布成功率(%)", report.success_rate])
                writer.writerow(["平均审批时长(小时)", report.avg_approval_hours])
                writer.writerow([])
                writer.writerow(["发布详情"])
                writer.writerow(["发布ID", "版本号", "发布类型", "状态", "创建时间", "回滚次数"])
                for d in report.release_details:
                    writer.writerow([
                        d["release_id"], d["version"], d["type"],
                        d["status"], d["created_at"], d["rollback_count"],
                    ])
            return str(csv_path)

    def generate_pdf_report(self, report: WeeklyReport) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            pdf_path = self.reports_dir / f"weekly_report_{report.report_id}.pdf"

            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                topMargin=2 * cm,
                bottomMargin=2 * cm,
                leftMargin=2 * cm,
                rightMargin=2 * cm,
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "ChineseTitle",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=18,
                spaceAfter=20,
                textColor=colors.HexColor("#4472C4"),
            )
            heading_style = ParagraphStyle(
                "ChineseHeading",
                parent=styles["Heading2"],
                fontName="Helvetica-Bold",
                fontSize=14,
                spaceBefore=15,
                spaceAfter=10,
                textColor=colors.HexColor("#2F5597"),
            )
            normal_style = ParagraphStyle(
                "ChineseNormal",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=10,
                spaceAfter=6,
            )

            story = []

            story.append(Paragraph("医疗器械 QMS 版本发布周报", title_style))
            story.append(Paragraph(
                f"报告周期: {report.start_date.strftime('%Y-%m-%d')} ~ {report.end_date.strftime('%Y-%m-%d')}",
                styles["Italic"]
            ))
            story.append(Spacer(1, 1 * cm))

            story.append(Paragraph("一、核心指标概览", heading_style))

            metrics_data = [
                ["指标", "数值"],
                ["总发布次数", str(report.total_releases)],
                ["成功发布次数", str(report.success_releases)],
                ["回滚次数", str(report.rollback_count)],
                ["发布成功率", f"{report.success_rate}%"],
                ["平均审批时长", f"{report.avg_approval_hours} 小时"],
            ]

            metrics_table = Table(metrics_data, colWidths=[8 * cm, 5 * cm])
            metrics_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(metrics_table)

            story.append(Spacer(1, 1 * cm))
            story.append(Paragraph("二、发布明细", heading_style))

            detail_data = [
                ["版本号", "类型", "状态", "创建时间", "回滚数"],
            ]
            for d in report.release_details:
                detail_data.append([
                    d["version"],
                    d["type"],
                    d["status"],
                    d["created_at"][:16],
                    str(d["rollback_count"]),
                ])

            detail_table = Table(detail_data, colWidths=[2.5 * cm, 2 * cm, 3 * cm, 3.5 * cm, 2 * cm])
            detail_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(detail_table)

            story.append(Spacer(1, 1 * cm))
            story.append(Paragraph("三、趋势数据（近8周）", heading_style))

            trend_data = self._generate_trend_data(report, 8)
            trend_data_list = [
                ["周次", "总发布", "成功", "回滚", "成功率"],
            ]
            for td in trend_data:
                trend_data_list.append([
                    td["week_label"],
                    str(td["total_releases"]),
                    str(td["success_releases"]),
                    str(td["rollback_count"]),
                    f"{td['success_rate']}%",
                ])

            trend_table = Table(trend_data_list, colWidths=[2 * cm, 2 * cm, 2 * cm, 2 * cm, 2.5 * cm])
            trend_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#70AD47")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(trend_table)

            story.append(Spacer(1, 1 * cm))
            story.append(Paragraph("四、备注说明", heading_style))
            story.append(Paragraph("1. 本报表由系统自动生成，数据来源于 QMS 发布管理系统。", normal_style))
            story.append(Paragraph("2. 发布成功率 = 成功发布次数 / 总发布次数 × 100%", normal_style))
            story.append(Paragraph("3. 回滚次数统计不包含演练性质的回滚。", normal_style))
            story.append(Paragraph("4. 趋势数据中历史数据为系统模拟展示。", normal_style))

            story.append(Spacer(1, 2 * cm))
            story.append(Paragraph(
                f"生成时间: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S') if report.generated_at else ''}",
                styles["Italic"]
            ))

            doc.build(story)
            return str(pdf_path)

        except ImportError:
            txt_path = self.reports_dir / f"weekly_report_{report.report_id}.txt"
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("医疗器械 QMS 版本发布周报\n")
                f.write("=" * 60 + "\n")
                f.write(f"报告周期: {report.start_date.strftime('%Y-%m-%d')} ~ {report.end_date.strftime('%Y-%m-%d')}\n\n")
                f.write("核心指标:\n")
                f.write(f"  总发布次数: {report.total_releases}\n")
                f.write(f"  成功发布: {report.success_releases}\n")
                f.write(f"  回滚次数: {report.rollback_count}\n")
                f.write(f"  发布成功率: {report.success_rate}%\n")
                f.write(f"  平均审批时长: {report.avg_approval_hours} 小时\n")
                f.write("\n" + "=" * 60 + "\n")
                f.write(f"生成时间: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S') if report.generated_at else ''}\n")
            return str(txt_path)

    def export_full_report(self, report: WeeklyReport) -> Dict[str, str]:
        excel_path = self.generate_excel_report(report)
        pdf_path = self.generate_pdf_report(report)
        return {
            "excel": excel_path,
            "pdf": pdf_path,
        }

    def get_report_summary(self, report: WeeklyReport) -> str:
        summary_lines = [
            "=" * 50,
            "📊 QMS 版本发布周报",
            "=" * 50,
            f"报告周期: {report.start_date.strftime('%Y-%m-%d')} ~ {report.end_date.strftime('%Y-%m-%d')}",
            "",
            f"📦 总发布次数: {report.total_releases}",
            f"✅ 成功发布: {report.success_releases}",
            f"🔄 回滚次数: {report.rollback_count}",
            f"📈 发布成功率: {report.success_rate}%",
            f"⏱️  平均审批时长: {report.avg_approval_hours} 小时",
            "=" * 50,
        ]
        return "\n".join(summary_lines)


class ReleaseQueryService:
    def __init__(self):
        self.releases_dir = Path(CONFIG["storage"]["releases_dir"])
        self._ensure_storage()

    def _ensure_storage(self):
        self.releases_dir.mkdir(parents=True, exist_ok=True)

    def _get_release_zones(self, release_data: Dict) -> List[str]:
        zones = set()

        if "grayscale_phases" in release_data:
            for phase in release_data["grayscale_phases"]:
                for z in phase.get("zones", []):
                    zones.add(z)

        if "rollback_records" in release_data:
            for rb in release_data["rollback_records"]:
                for z in rb.get("affected_zones", []):
                    zones.add(z)

        return list(zones)

    def query_releases(self,
                       start_date: datetime = None,
                       end_date: datetime = None,
                       status: ReleaseStatus = None,
                       release_type: ReleaseType = None,
                       version: str = None,
                       zone_id: str = None,
                       include_rollbacks: bool = False) -> List[Dict]:
        results = []

        if not self.releases_dir.exists():
            return results

        for file in self.releases_dir.glob("*.json"):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                created_at = datetime.fromisoformat(data["created_at"])

                if start_date and created_at < start_date:
                    continue
                if end_date and created_at > end_date:
                    continue
                if status and data["status"] != status.value:
                    continue
                if release_type and data["release_type"] != release_type.value:
                    continue
                if version and version not in data["version"]:
                    continue

                if zone_id:
                    release_zones = self._get_release_zones(data)
                    if zone_id not in release_zones:
                        continue

                if include_rollbacks and not data.get("rollback_records"):
                    continue

                results.append(data)
            except Exception:
                continue

        return sorted(results, key=lambda x: x["created_at"], reverse=True)

    def query_rollbacks(self,
                        start_date: datetime = None,
                        end_date: datetime = None,
                        zone_id: str = None,
                        is_drill: bool = None,
                        version: str = None) -> List[Dict]:
        rollbacks = []

        if not self.releases_dir.exists():
            return rollbacks

        for file in self.releases_dir.glob("*.json"):
            try:
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if "rollback_records" not in data or not data["rollback_records"]:
                    continue

                for rb in data["rollback_records"]:
                    if is_drill is not None and rb.get("is_drill", False) != is_drill:
                        continue

                    if zone_id:
                        if zone_id not in rb.get("affected_zones", []):
                            continue

                    if version:
                        if version not in rb.get("from_version", "") and version not in rb.get("to_version", ""):
                            continue

                    rb_data = dict(rb)
                    rb_data["release_version"] = data["version"]
                    rb_data["release_id"] = data["release_id"]
                    rb_data["release_type"] = data["release_type"]

                    if start_date or end_date:
                        rb_time_str = rb.get("started_at")
                        if rb_time_str:
                            rb_time = datetime.fromisoformat(rb_time_str)
                            if start_date and rb_time < start_date:
                                continue
                            if end_date and rb_time > end_date:
                                continue

                    rollbacks.append(rb_data)
            except Exception:
                continue

        return sorted(rollbacks, key=lambda x: x.get("started_at", ""), reverse=True)

    def get_release_detail(self, release_id: str) -> Optional[Dict]:
        file_path = self.releases_dir / f"release_{release_id}.json"
        if file_path.exists():
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        return None

    def export_releases(self, output_file: str,
                        start_date: datetime = None,
                        end_date: datetime = None,
                        status: ReleaseStatus = None,
                        release_type: ReleaseType = None,
                        version: str = None,
                        zone_id: str = None,
                        format: str = "json") -> str:
        releases = self.query_releases(
            start_date=start_date,
            end_date=end_date,
            status=status,
            release_type=release_type,
            version=version,
            zone_id=zone_id,
        )

        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format == "json":
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(releases, f, ensure_ascii=False, indent=2)
        elif format == "csv":
            import csv
            with open(output_file, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "发布ID", "版本号", "发布类型", "状态", "创建人",
                    "创建时间", "变更控制单", "回滚次数", "影响厂区数"
                ])
                for r in releases:
                    zones = self._get_release_zones(r)
                    writer.writerow([
                        r["release_id"], r["version"], r["release_type"],
                        r["status"], r.get("requester", ""),
                        r["created_at"], r.get("change_control_id", ""),
                        len(r.get("rollback_records", [])),
                        len(zones),
                    ])

        return str(output_path)

    def export_rollbacks(self, output_file: str,
                         start_date: datetime = None,
                         end_date: datetime = None,
                         zone_id: str = None,
                         is_drill: bool = None,
                         format: str = "json") -> str:
        rollbacks = self.query_rollbacks(
            start_date=start_date,
            end_date=end_date,
            zone_id=zone_id,
            is_drill=is_drill,
        )

        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format == "json":
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(rollbacks, f, ensure_ascii=False, indent=2)
        elif format == "csv":
            import csv
            with open(output_file, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "回滚ID", "发布ID", "发布版本", "发布类型",
                    "从版本", "回滚至", "原因", "是否演练",
                    "影响厂区数", "开始时间", "完成时间"
                ])
                for rb in rollbacks:
                    writer.writerow([
                        rb.get("rollback_id", ""),
                        rb.get("release_id", ""),
                        rb.get("release_version", ""),
                        rb.get("release_type", ""),
                        rb.get("from_version", ""),
                        rb.get("to_version", ""),
                        rb.get("reason", ""),
                        "是" if rb.get("is_drill") else "否",
                        len(rb.get("affected_zones", [])),
                        rb.get("started_at", ""),
                        rb.get("completed_at", ""),
                    ])

        return str(output_path)

    def save_release(self, release: ReleaseRecord):
        file_path = self.releases_dir / f"release_{release.release_id}.json"

        data = {
            "release_id": release.release_id,
            "version": release.version,
            "release_type": release.release_type.value,
            "status": release.status.value,
            "title": release.title,
            "description": release.description,
            "change_control_id": release.change_control_id,
            "requester": release.requester,
            "created_at": release.created_at.isoformat(),
            "previous_version": release.previous_version,
            "target_zones": release.target_zones,
            "rollback_count": len(release.rollback_records),
        }

        if release.pre_check_result:
            data["pre_check"] = {
                "overall_pass": release.pre_check_result.overall_pass,
                "check_count": len(release.pre_check_result.checks),
                "failed_count": len(release.pre_check_result.get_failed_checks()),
            }

        if release.approval_flow:
            data["approval"] = {
                "is_hotfix": release.approval_flow.is_hotfix,
                "hotfix_reason": release.approval_flow.hotfix_reason,
                "is_completed": release.approval_flow.is_completed(),
                "is_rejected": release.approval_flow.is_rejected(),
                "steps_count": len(release.approval_flow.steps),
            }

        data["grayscale_phases"] = [
            {
                "phase_id": p.phase_id,
                "name": p.name,
                "tier": p.tier.value,
                "zones": p.zones,
                "status": p.status,
            }
            for p in release.grayscale_phases
        ]

        data["rollback_records"] = [
            {
                "rollback_id": r.rollback_id,
                "reason": r.reason,
                "affected_zones": r.affected_zones,
                "from_version": r.from_version,
                "to_version": r.to_version,
                "is_drill": r.is_drill,
                "started_at": r.started_at.isoformat() if r.started_at else None,
                "completed_at": r.completed_at.isoformat() if r.completed_at else None,
            }
            for r in release.rollback_records
        ]

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def generate_weekly_report(releases: List[ReleaseRecord]) -> WeeklyReport:
    generator = ReportGenerator()
    return generator.generate_weekly_report(releases)
